import openpyxl as xl

#Type in the name of file
enter_work_book = input("Enter work book: ")

#Concatinate it with the ".xlsx" extension, to save it as excel file
file_name = enter_work_book + ".xlsx"

#Load a new workbook using "xl.load_workbook()" and then added the "file_name" from above
wb = xl.load_workbook(file_name)

#Load another file where we are importing the data to be used from
wb2 = xl.load_workbook("org_samples.xlsx")



def automate_spread_sheet(input_file,import_data_from_file):
    sheets2 = import_data_from_file["Sheet"]
    sheets = input_file["Sheet1"]

    products_tab = sheets["A1"]
    products_tab.value = "Products"
    transaction_tab = sheets["B1"]
    transaction_tab.value = "Transactions_Id"
    price_tab = sheets["C1"]
    price_tab.value = "New Price"
    previous_price = sheets["D1"]
    previous_price.value = "Previous Price"

    for row in range(2,sheets2.max_row + 1):
        prod_cell = sheets.cell(row,1)
        trans_cell = sheets.cell(row,2)
        price_cell = sheets.cell(row,3)
        previous_price_cell = sheets.cell(row,4)

        prod_cell.value = sheets2.cell(row,4).value
        price_cell.value = sheets2.cell(row,3).value
        trans_cell.value = sheets2.cell(row,1).value
        previous_price_cell.value = price_cell.value * 1.24

        print(previous_price.value, previous_price_cell.value)

        wb.save(file_name)



automate_spread_sheet(wb, wb2)
